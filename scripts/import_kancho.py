#!/usr/bin/env python3
# 班長シフト（管理者公休予定表）Excelインポート
# 「管理者公休予定表　2025年～.xlsx」の指定シートからメイン表（班長シフト）と
# 特記事項・希望休メモを読み取り、D1投入用SQLを生成する。
#
# 使い方:
#   python3 scripts/import_kancho.py                # → scripts/import_kancho.sql を生成
#   cd system && npx wrangler d1 execute staff-db --remote --file=../scripts/import_kancho.sql
#
# 仕様メモ:
# - シートは「2026.6」のように後ろの月が月度。日付は前月10日ごろ〜当月18日ごろ。
# - 「中島→船崎」「誠⇒矢嶋」のような交代表記は矢印の後ろ（現任者）の名前で登録する。
# - Excelの書式も取り込む:
#     ・セル塗り＝班色（黄緑/黄色/水色/ピンク）。本人の班色はセルの最頻色から推定
#     ・本人の班色と違う塗りのセル（他班ヘルプ等）は cell_color として個別保存
#     ・斜体の「直」＝斜め直 → is_diagonal=1
#     ・赤文字＝希望休の反映 → is_wish=1
#     ・空白セルでも班色と違う塗りがあれば code='' + cell_color で保存
# - 下段の①②表・前川/前田の小表は日付ヘッダーが古い月のまま残っているため取り込まない。
# - 再実行しても安全（シフトは INSERT OR REPLACE、メモは月度ごとに DELETE→INSERT）。

import datetime
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl

EXCEL = Path(__file__).parent.parent / '管理者公休予定表　2025年～.xlsx'
OUT = Path(__file__).parent / 'import_kancho.sql'

# 取り込むシート（時系列順。後のシートが重複日付を上書きする）
SHEETS = ['2026.6', '2026.7(事故教育加入ver)', '2026.8']

ROLES = ['昼日勤班長', '終業班長', '教育班長', '研修課出向', '職員当直']
MEMBER_ROW_START = 8

# Excelの塗り色 → Web用カラーコード
FILL_MAP = {
    'FF00FF00': '#00ff00',  # 黄緑
    'FFFFFF00': '#ffff00',  # 黄色
    'FF00FFFF': '#00ffff',  # 水色
    'FFFF99CC': '#ff99cc',  # ピンク
    'FFFF0000': '#ff0000',  # 赤
    'FFA5A5A5': '#a5a5a5',  # グレー
}
TEAM_COLORS = {'#00ff00', '#ffff00', '#00ffff', '#ff99cc'}  # 班色として推定対象


def q(s: str) -> str:
    return s.replace("'", "''")


def norm_name(raw) -> str:
    """「中島→船崎」→「船崎」。空白除去。"""
    s = str(raw).strip().replace('　', '').replace(' ', '')
    s = re.split('[→⇒]', s)[-1]
    return s.strip()


def cell_text(v) -> str:
    if v is None:
        return ''
    if isinstance(v, datetime.datetime):
        return f'{v.month}/{v.day}'
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    if isinstance(v, int) and 40000 < v < 60000:  # Excelシリアル日付が生で入っている場合
        d = datetime.date(1899, 12, 30) + datetime.timedelta(days=v)
        return f'{d.month}/{d.day}'
    return str(v).strip()


def cell_fill(cell):
    f = cell.fill
    if f and f.fill_type == 'solid':
        rgb = f.start_color.rgb
        if isinstance(rgb, str):
            return FILL_MAP.get(rgb)
    return None


def is_red_font(cell) -> bool:
    return bool(cell.font and cell.font.color and cell.font.color.rgb == 'FFFF0000')


def day_of(v):
    if isinstance(v, datetime.datetime):
        return v.day
    if isinstance(v, (int, float)) and 1 <= v <= 31:
        return int(v)
    return None


def parse_sheet(ws, sheet_name: str):
    """(dates_by_col, period_year, period_month) を返す"""
    m = re.match(r'^(\d{4})\.(\d{1,2})', sheet_name)
    if not m:
        raise ValueError(f'シート名から年月が読めません: {sheet_name}')
    year2, month2 = int(m.group(1)), int(m.group(2))  # 後ろの月＝月度
    month1 = month2 - 1 if month2 > 1 else 12
    year1 = year2 if month2 > 1 else year2 - 1

    dates_by_col = {}
    cur_month, cur_year = month1, year1
    prev_day = 0
    for col in range(3, ws.max_column + 1):
        d = day_of(ws.cell(7, col).value)
        if d is None:
            break
        if d < prev_day:  # 月替わり
            cur_month = month2
            cur_year = year2
        prev_day = d
        dates_by_col[col] = f'{cur_year:04d}-{cur_month:02d}-{d:02d}'
    if not dates_by_col:
        raise ValueError(f'{sheet_name}: 日付ヘッダー（7行目）が見つかりません')
    return dates_by_col, year2, month2


def find_marker(ws, text: str):
    for r in range(1, 35):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip() == text:
                return r, c
    return None


def member_rows(ws):
    """(row, name, role) のリスト。名前のない行はスキップ"""
    out = []
    row = MEMBER_ROW_START
    while True:
        role = ws.cell(row, 2).value
        if not isinstance(role, str) or role.strip() not in ROLES:
            break
        name_raw = ws.cell(row, 1).value
        if name_raw:
            name = norm_name(name_raw)
            if name:
                out.append((row, name, role.strip()))
        row += 1
    return out


def main():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)  # data_onlyでも書式は取得できる
    sql = [
        '-- 班長シフト Excelインポート（自動生成: scripts/import_kancho.py）',
        f"-- 生成日時: {datetime.datetime.now().isoformat(timespec='seconds')}",
        f"-- 対象シート: {', '.join(SHEETS)}",
        '',
    ]
    member_order = {}   # name -> (role, sort_order) 最後に出たシートの値
    member_color = {}   # name -> 班色（最後に出たシートの最頻色）
    last_sheet_names = set()  # 最終シートに載っている名前 = 現役の内勤班長
    shift_count = 0

    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f'警告: シート {sheet_name} が見つかりません。スキップします', file=sys.stderr)
            continue
        ws = wb[sheet_name]
        dates_by_col, pyear, pmonth = parse_sheet(ws, sheet_name)
        first = list(dates_by_col.values())[0]
        last = list(dates_by_col.values())[-1]
        sql.append(f'-- ===== {sheet_name}（{first} 〜 {last}） =====')

        rows = member_rows(ws)
        last_sheet_names = {name for _r, name, _ro in rows}  # ループ最後のシートの値が残る

        # 1周目: 各人の班色を「空白セルの塗り」の最頻色から推定
        # （昼日勤班長だけが空白=出勤も班色で塗られている。終業・教育班長は空白が無塗り=班色なし）
        sheet_color = {}
        for row, name, _role in rows:
            fills = Counter()
            for col in dates_by_col:
                cell = ws.cell(row, col)
                if cell.value is None:
                    fl = cell_fill(cell)
                    if fl in TEAM_COLORS:
                        fills[fl] += 1
            if fills and fills.most_common(1)[0][1] >= 3:
                sheet_color[name] = fills.most_common(1)[0][0]
            # このシートでの推定結果で常に上書き（役割が変わって班色が外れた場合も追従）
            member_color[name] = sheet_color.get(name)

        # 2周目: シフトセルを書式込みで出力
        sort_order = 10
        for row, name, role in rows:
            member_order[name] = (role, sort_order)
            sort_order += 10
            my_color = sheet_color.get(name)
            sql.append(
                f"INSERT INTO kancho_members (name, role, section, sort_order) "
                f"SELECT '{q(name)}', '{q(role)}', 'main', {member_order[name][1]} "
                f"WHERE NOT EXISTS (SELECT 1 FROM kancho_members WHERE name = '{q(name)}' AND section = 'main');"
            )
            for col, date in dates_by_col.items():
                cell = ws.cell(row, col)
                code = cell_text(cell.value)
                fl = cell_fill(cell)
                dg = 1 if (cell.font and cell.font.italic and code) else 0
                wish = 1 if is_red_font(cell) else 0
                # 本人の班色と違う塗りだけ個別色として保存（同じ塗りは自動表示に任せる）
                cl = fl if (fl and fl != my_color) else None
                if not code and not cl:
                    continue  # 完全な空白（班色 or 塗りなし）= 自動表示
                cl_sql = f"'{cl}'" if cl else 'NULL'
                sql.append(
                    f"INSERT OR REPLACE INTO kancho_shifts (member_id, date, code, is_diagonal, is_wish, cell_color, updated_by) "
                    f"SELECT id, '{date}', '{q(code)}', {dg}, {wish}, {cl_sql}, 'excel-import' FROM kancho_members "
                    f"WHERE name = '{q(name)}' AND section = 'main';"
                )
                shift_count += 1

        # メモ（特記事項・希望休）→ 月度に紐付け。再実行に備えて先に削除
        memo_stmts = []
        mk = find_marker(ws, '・特記事項')
        if mk:
            r, c = mk
            lines = []
            for rr in range(r + 1, r + 6):
                v = ws.cell(rr, c).value
                if v is None or (isinstance(v, str) and not v.strip()):
                    break
                lines.append(cell_text(v))
            if lines:
                memo_stmts.append(
                    f"INSERT INTO kancho_memos (year, month, kind, title, content, sort_order) "
                    f"VALUES ({pyear}, {pmonth}, 'tokki', '', '{q(chr(10).join(lines))}', 0);"
                )
        mk = find_marker(ws, '・希望休')
        if mk:
            r, c = mk
            order = 10
            for rr in range(r + 1, r + 15):
                name_v = ws.cell(rr, c).value
                if name_v is None or (isinstance(name_v, str) and not name_v.strip()):
                    break
                content = cell_text(ws.cell(rr, c + 2).value)
                memo_stmts.append(
                    f"INSERT INTO kancho_memos (year, month, kind, title, content, sort_order) "
                    f"VALUES ({pyear}, {pmonth}, 'kibou', '{q(cell_text(name_v))}', '{q(content)}', {order});"
                )
                order += 10
        if memo_stmts:
            sql.append(f'DELETE FROM kancho_memos WHERE year = {pyear} AND month = {pmonth};')
            sql.extend(memo_stmts)
        sql.append('')

    # 最終シートの並び・役割・班色でメンバーを更新
    # 最終シートに載っていない人 = 交代して乗務に戻った班長 → 内勤オフ（表に非表示）
    sql.append('-- 最新シートの役割・並び順・班色・内勤フラグを反映')
    for name, (role, order) in member_order.items():
        color = member_color.get(name)
        color_sql = f"'{color}'" if color else 'NULL'
        indoor = 1 if name in last_sheet_names else 0
        sql.append(
            f"UPDATE kancho_members SET role = '{q(role)}', sort_order = {order}, team_color = {color_sql}, is_indoor = {indoor} "
            f"WHERE name = '{q(name)}' AND section = 'main';"
        )

    OUT.write_text('\n'.join(sql) + '\n', encoding='utf-8')
    print(f'生成完了: {OUT}')
    print(f'  メンバー: {len(member_order)}名 / シフト: {shift_count}件')
    print('  班色推定:', {k: v for k, v in member_color.items()})


if __name__ == '__main__':
    main()
