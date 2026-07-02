#!/usr/bin/env python3
"""
Excelシフトデータ → Cloudflare D1 インポートスクリプト
対象: こぴー2023年度 新人研修日数 2026.06作成②.xlsx
移行範囲: 5月度以降（2026.05月度、2026.06月度...）

使い方:
  pip install openpyxl
  python3 import_excel.py --file "/path/to/excel.xlsx" --output import.sql
  wrangler d1 execute staff-db --file=import.sql
  # または本番に直接投入:
  wrangler d1 execute staff-db --remote --file=import.sql
"""

import argparse
import re
from datetime import datetime, date, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxlをインストールしてください: pip install openpyxl")
    raise

# ============================================================
# 設定
# ============================================================
TARGET_SHEETS = ['2026.05月度', '2026.06月度', '2026.07月度']

# 列インデックス（0始まり）
COL_SEQ_NO = 0    # NO
COL_DIVISION = 1  # 課
COL_TEAM = 2      # 班
COL_HIRE_DATE = 3 # 配属日
COL_ENTRY_TYPE = 4 # 初乗務（新卒/縁故/キャリア等）
COL_LOCKER = 5    # ロッカー番号
COL_EMP_NO = 6    # 社員番号
COL_NAME = 7      # 氏名
COL_NAME_KANA = 8 # カナ氏名
COL_DATE_START = 9  # 日付列の開始

# 月と年の対応（シート名から判断）
# 例: "2026.06月度" → year=2026, month=6
def parse_sheet_name(name: str):
    m = re.match(r'(\d{4})\.(\d{2})月度', name)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None

def safe_str(val) -> str:
    if val is None:
        return ''
    return str(val).strip()

def safe_int(val):
    if val is None:
        return None
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return None

def escape_sql(s: str) -> str:
    return s.replace("'", "''")

def sql_val(v) -> str:
    if v is None:
        return 'NULL'
    if isinstance(v, (int, float)):
        return str(int(v))
    return f"'{escape_sql(str(v))}'"

def parse_date(val):
    """配属日など日付を YYYY-MM-DD 形式に変換"""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        if isinstance(val, datetime):
            return val.strftime('%Y-%m-%d')
        return val.isoformat()
    # 数値（Excelシリアル値）の場合
    try:
        n = int(float(str(val)))
        # Excel日付シリアル値 → Python datetime
        excel_epoch = datetime(1899, 12, 30)
        d = excel_epoch + timedelta(days=n)
        return d.strftime('%Y-%m-%d')
    except (ValueError, TypeError):
        pass
    return None

def determine_entry_type(val) -> str:
    """初乗務列の値からentry_typeを判断"""
    s = safe_str(val).lower()
    if '新卒' in s:
        return '新卒'
    if '縁故' in s:
        return '縁故'
    if 'キャリア' in s or '中途' in s:
        return 'キャリア'
    # 日付が入っている場合は初乗務日（entry_typeは不明→新卒とする）
    return 'キャリア'  # デフォルト：既存社員はキャリア扱い

def read_sheet(ws) -> dict:
    """1シートを読み込んで社員データとシフトデータを返す"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {'employees': [], 'shifts': []}

    # ヘッダー行（NO, 課, 班...の行）を探す
    header_row_idx = None
    date_row_idx = None
    for i, row in enumerate(rows):
        if row[0] == 'NO' or (row[0] is not None and str(row[0]).strip() == 'NO'):
            header_row_idx = i
            date_row_idx = i  # 同じ行に日付も入っている
            break

    if header_row_idx is None:
        # 日付が数値で入っている行を探す
        for i, row in enumerate(rows):
            if len(row) > COL_DATE_START and isinstance(row[COL_DATE_START], (int, float)):
                header_row_idx = i
                break

    if header_row_idx is None:
        return {'employees': [], 'shifts': []}

    # 日付列の実際の日付を特定
    # 3行目(インデックス2)が日付番号行、4行目が曜日行
    # 実際の列: COL_DATE_START以降が日付列
    date_row = rows[header_row_idx]
    dates_in_columns = {}  # col_idx → date string

    # シート名から年月を取得してYYYY-MM-DDを組み立てる
    # 日付番号（1〜31）と月度から実際の日付を決定する
    sheet_title = ws.title
    sheet_year, sheet_month = parse_sheet_name(sheet_title)

    if sheet_year and sheet_month:
        # 月度: 前月18日〜当月17日
        # 前月
        prev_month = sheet_month - 1
        prev_year = sheet_year
        if prev_month < 1:
            prev_month = 12
            prev_year -= 1

        for col_i in range(COL_DATE_START, len(date_row)):
            day_num = date_row[col_i]
            if day_num is None:
                continue
            try:
                day = int(float(str(day_num)))
                if 1 <= day <= 31:
                    # 15〜31 → 前月, 1〜17 → 当月
                    if day >= 15:
                        try:
                            d = date(prev_year, prev_month, day)
                            dates_in_columns[col_i] = d.isoformat()
                        except ValueError:
                            pass
                    else:
                        try:
                            d = date(sheet_year, sheet_month, day)
                            dates_in_columns[col_i] = d.isoformat()
                        except ValueError:
                            pass
            except (ValueError, TypeError):
                pass

    # データ行を処理（ヘッダー行の次から）
    employees = {}  # emp_no → employee dict
    shifts = []     # (emp_no, date, entry_main, entry_sub) のリスト

    data_start = header_row_idx + 2  # ヘッダー + 曜日行をスキップ

    i = data_start
    while i < len(rows):
        row = rows[i]
        emp_no = safe_str(row[COL_EMP_NO] if len(row) > COL_EMP_NO else None)

        # 有効な社員番号行かチェック（数値7-8桁）
        if re.match(r'^\d{7,8}$', emp_no):
            name = safe_str(row[COL_NAME] if len(row) > COL_NAME else None).replace('　', ' ')
            name_kana = safe_str(row[COL_NAME_KANA] if len(row) > COL_NAME_KANA else None).replace('　', ' ')
            division = safe_int(row[COL_DIVISION] if len(row) > COL_DIVISION else None)
            team = safe_int(row[COL_TEAM] if len(row) > COL_TEAM else None)
            locker = safe_str(row[COL_LOCKER] if len(row) > COL_LOCKER else None)
            seq_no = safe_int(row[COL_SEQ_NO] if len(row) > COL_SEQ_NO else None)
            hire_date = parse_date(row[COL_HIRE_DATE] if len(row) > COL_HIRE_DATE else None)
            entry_type_raw = row[COL_ENTRY_TYPE] if len(row) > COL_ENTRY_TYPE else None
            entry_type = determine_entry_type(entry_type_raw)

            # 社員番号が2026以降ならキャリア年度から判断
            emp_no_year = emp_no[:4] if len(emp_no) >= 4 else ''
            if emp_no[:4] == '2026' and len(emp_no) == 8:
                entry_type_str = safe_str(entry_type_raw)
                if '新卒' in entry_type_str:
                    entry_type = '新卒'
                elif '縁故' in entry_type_str:
                    entry_type = '縁故'
                else:
                    entry_type = '新卒'  # 2026年入社はデフォルト新卒
            elif emp_no[:4] in ('2025', '2024', '2023'):
                entry_type = 'キャリア'

            if emp_no not in employees:
                employees[emp_no] = {
                    'emp_no': emp_no,
                    'name': name,
                    'name_kana': name_kana,
                    'division': division,
                    'team': team,
                    'locker_no': locker or None,
                    'seq_no': seq_no,
                    'hire_date': hire_date,
                    'entry_type': entry_type,
                }

            # メイン行のシフト（dates_in_columnsのキーを使って）
            for col_i, date_str in dates_in_columns.items():
                if col_i < len(row):
                    entry_main = safe_str(row[col_i])
                    # 次の行（詳細行）を読む
                    entry_sub = ''
                    if i + 1 < len(rows):
                        next_row = rows[i + 1]
                        if (len(next_row) > COL_EMP_NO and
                            (next_row[COL_EMP_NO] is None or safe_str(next_row[COL_EMP_NO]) == '') and
                            col_i < len(next_row)):
                            entry_sub = safe_str(next_row[col_i])

                    if entry_main or entry_sub:
                        shifts.append((emp_no, date_str, entry_main or None, entry_sub or None))

        i += 1

    return {
        'employees': list(employees.values()),
        'shifts': shifts
    }

def generate_sql(all_employees: dict, all_shifts: list) -> str:
    lines = [
        '-- ========================================',
        '-- Excelデータ移行SQL (2026.05月度以降)',
        f'-- 生成日時: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
        '-- ========================================',
        '',
        '-- 社員データ',
    ]

    for emp in all_employees.values():
        lines.append(
            f"INSERT OR IGNORE INTO employees "
            f"(emp_no, name, name_kana, division, team, locker_no, seq_no, hire_date, entry_type) VALUES ("
            f"{sql_val(emp['emp_no'])}, "
            f"{sql_val(emp['name'])}, "
            f"{sql_val(emp['name_kana']) if emp['name_kana'] else 'NULL'}, "
            f"{sql_val(emp['division'])}, "
            f"{sql_val(emp['team'])}, "
            f"{sql_val(emp['locker_no'])}, "
            f"{sql_val(emp['seq_no'])}, "
            f"{sql_val(emp['hire_date'])}, "
            f"{sql_val(emp['entry_type'])}"
            f");"
        )

    lines += ['', '-- シフトデータ']
    for emp_no, shift_date, entry_main, entry_sub in all_shifts:
        lines.append(
            f"INSERT OR IGNORE INTO shift_entries "
            f"(emp_id, date, entry_main, entry_sub) "
            f"SELECT id, {sql_val(shift_date)}, {sql_val(entry_main)}, {sql_val(entry_sub)} "
            f"FROM employees WHERE emp_no = {sql_val(emp_no)};"
        )

    lines += ['', f'-- 合計: 社員 {len(all_employees)}名, シフト {len(all_shifts)}件']
    return '\n'.join(lines)

def main():
    parser = argparse.ArgumentParser(description='ExcelシフトデータをD1用SQLに変換')
    parser.add_argument('--file', required=True, help='Excelファイルのパス')
    parser.add_argument('--output', default='import.sql', help='出力SQLファイル名')
    parser.add_argument('--sheets', nargs='+', default=TARGET_SHEETS, help='処理するシート名')
    args = parser.parse_args()

    print(f"読み込み中: {args.file}")
    wb = openpyxl.load_workbook(args.file, data_only=True)

    all_employees = {}
    all_shifts = []

    for sheet_name in args.sheets:
        if sheet_name not in wb.sheetnames:
            print(f"  スキップ（存在しない）: {sheet_name}")
            continue
        print(f"  処理中: {sheet_name}")
        ws = wb[sheet_name]
        result = read_sheet(ws)
        for emp in result['employees']:
            key = emp['emp_no']
            if key not in all_employees:
                all_employees[key] = emp
            else:
                # より新しい情報で更新（課・班・ロッカー等）
                for field in ('division', 'team', 'locker_no', 'hire_date'):
                    if emp[field] and not all_employees[key][field]:
                        all_employees[key][field] = emp[field]
        all_shifts.extend(result['shifts'])

    # 重複シフトを除去（同一emp_no + dateで最後のものを採用）
    shift_map = {}
    for s in all_shifts:
        key = f"{s[0]}_{s[1]}"
        shift_map[key] = s
    unique_shifts = list(shift_map.values())

    sql = generate_sql(all_employees, unique_shifts)
    output_path = Path(args.output)
    output_path.write_text(sql, encoding='utf-8')

    print(f"\n✅ 完了!")
    print(f"  社員数: {len(all_employees)}名")
    print(f"  シフト数: {len(unique_shifts)}件")
    print(f"  出力ファイル: {output_path.absolute()}")
    print(f"\n次のコマンドでD1にインポート:")
    print(f"  # ローカルテスト:")
    print(f"  wrangler d1 execute staff-db --local --file={args.output}")
    print(f"  # 本番投入:")
    print(f"  wrangler d1 execute staff-db --remote --file={args.output}")

if __name__ == '__main__':
    main()
