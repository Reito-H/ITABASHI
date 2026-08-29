# ブラウザ側 xfSetText/xfSetNum と同じ「<c> の中身だけ差し替え（s= 保持）」を Python で再現し、
# 記入済みサンプルを生成する。デザインが崩れないか openpyxl で読み直して検証＋プレビュー。
import re, io, zipfile, shutil

def xf_set_cell(xml, ref, inner, extra=''):
    m = re.search(r'<c r="%s"' % re.escape(ref), xml)
    if not m:
        raise SystemExit('cell %s not found' % ref)
    i = m.start()
    gt = xml.index('>', i)
    head = xml[i:gt]
    self_close = xml[gt-1] == '/'
    end = gt+1 if self_close else xml.index('</c>', gt)+4
    sm = re.search(r' s="(\d+)"', head)
    s = ' s="%s"' % sm.group(1) if sm else ''
    rep = '<c r="%s"%s%s>%s</c>' % (ref, s, extra, inner)
    return xml[:i] + rep + xml[end:]

def esc(v):
    return (str(v).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')
            .replace('"','&quot;').replace("'",'&apos;'))

def set_text(xml, ref, text):
    if text == '' or text is None:
        return xf_set_cell(xml, ref, '', '')
    return xf_set_cell(xml, ref, '<is><t xml:space="preserve">%s</t></is>' % esc(text), ' t="inlineStr"')

def set_num(xml, ref, num):
    if num == '' or num is None:
        return xml
    return xf_set_cell(xml, ref, '<v>%s</v>' % num, '')

def fill(src, dst, edits):
    z = zipfile.ZipFile(src)
    names = z.namelist()
    data = {n: z.read(n) for n in names}
    s1 = data['xl/worksheets/sheet1.xml'].decode()
    for fn, ref, val in edits:
        s1 = (set_text if fn == 't' else set_num)(s1, ref, val)
    data['xl/worksheets/sheet1.xml'] = s1.encode()
    with zipfile.ZipFile(dst, 'w', zipfile.ZIP_DEFLATED) as o:
        for n in names:
            o.writestr(n, data[n])
    # 検証: openpyxl で開けるか & 結合数保持
    from openpyxl import load_workbook
    wb1 = load_workbook(src); wb2 = load_workbook(dst)
    m1 = sorted(str(m) for m in wb1.active.merged_cells.ranges)
    m2 = sorted(str(m) for m in wb2.active.merged_cells.ranges)
    assert m1 == m2, 'merge mismatch!'
    print(dst, 'OK  merges', len(m2))

# ---- 顛末書 ----
fill('tenmatsusho_template.xlsx', 'tenmatsusho_sample.xlsx', [
    ('t', 'A5', '原　義夫'),
    ('n', 'P2', 2026), ('n', 'S2', 8), ('n', 'U2', 29),
    ('n', 'P5', 2), ('n', 'T5', 3),
    ('t', 'P6', '田中　太郎'), ('t', 'P7', '00012345'),
    ('n', 'E13', 2026), ('n', 'H13', 8), ('n', 'J13', 27), ('t', 'M13', '水'),
    ('t', 'R13', '午後 3'), ('t', 'T13', '20'),
    ('t', 'E14', '首都高速 5号池袋線 下り 熊野町JCT付近'),
    ('t', 'E15', '追突（物損）'), ('t', 'S15', '板橋123'),
    ('t', 'A20', '渋滞最後尾で前車が急停止し、車間距離が不足していたため追突した。'),
    ('t', 'A32', '車間距離不保持。再発防止のため添乗指導を実施。'),
    ('t', 'S32', '原　義夫'), ('t', 'U32', '柴村　昌幸'),
])

# ---- 羽田 ----
fill('haneda_riyusho_template.xlsx', 'haneda_sample.xlsx', [
    ('t', 'H3', '板橋'),
    ('n', 'D5', 2), ('n', 'N5', 3),
    ('n', 'W5', 26), ('n', 'AB5', 8), ('n', 'AF5', 29),
    ('t', 'I7', '00012345'), ('t', 'AC7', '田中　太郎'),
    ('n', 'K8', 26), ('n', 'O8', 8), ('n', 'R8', 27), ('t', 'U8', '水'),
    ('t', 'AG8', '足立 500 あ 12-34'),
    ('t', 'J9', '14'), ('t', 'N9', '05'), ('t', 'J10', '15'), ('t', 'N10', '30'),
    ('t', 'AA9', '羽田空港　　T1 ・ 【T2】 ・ T3'),
    ('t', 'AA10', '東京都新宿区西新宿2-8-1'),
    ('n', 'AG11', 2),
    ('t', 'I11', '無線配車　・　【定額乗場】'),
    ('t', 'I12', '日本人　・　【外国人】　・　不明'),
    ('t', 'AG12', '可　・　【不可】'),
    ('t', 'C16', '■'), ('t', 'C18', '■'),
    ('t', 'W18', '大井町駅経由'),
    ('t', 'D20', '定額範囲外の経由地に立ち寄ったため、メーター運賃を適用した。'),
    ('t', 'AA29', '柴村　昌幸'), ('t', 'AH29', '原　義夫'),
])
print('done')
