# 羽田定額適用外理由書（国際自動車）.pdf を xlsx で再現する。
# PDF はフォームフィールドを持たないフラット帳票なので、罫線グリッドを引き直す。
import base64, io, zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.cell.cell import MergedCell

OUT_XLSX = 'haneda_riyusho_template.xlsx'
OUT_TS = 'haneda_riyusho_template.ts'

NCOL = 40
FONT = 'ＭＳ Ｐゴシック'
L = get_column_letter

wb = Workbook()
ws = wb.active
ws.title = '羽田定額適用外理由書'

for c in range(1, NCOL + 1):
    ws.column_dimensions[L(c)].width = 2.4

ROWH = {
    1: 32, 2: 10, 3: 22, 4: 8, 5: 22, 6: 10,
    7: 24, 8: 24, 9: 26, 10: 24, 11: 22, 12: 22,
    13: 16, 14: 26, 15: 10,
    16: 24, 17: 24, 18: 24, 19: 24,
    20: 22, 21: 22, 22: 22, 23: 22,
    24: 20, 25: 20, 26: 20, 27: 20,
    28: 19, 29: 46,
}
for r, h in ROWH.items():
    ws.row_dimensions[r].height = h

def S(style='thin'):
    return Side(style=style)

def edge(r1, c1, r2, c2, sides='TBLR', style='thin'):
    s = S(style)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            b = cell.border
            top, bot, lft, rgt = b.top, b.bottom, b.left, b.right
            if 'T' in sides and r == r1: top = s
            if 'B' in sides and r == r2: bot = s
            if 'L' in sides and c == c1: lft = s
            if 'R' in sides and c == c2: rgt = s
            cell.border = Border(top=top, bottom=bot, left=lft, right=rgt)

def grid(r1, c1, r2, c2):
    edge(r1, c1, r2, c2, 'TBLR')

def cell(ref, text='', size=10.5, bold=False, ha='left', va='center', wrap=False, to=None):
    cs, ri = coordinate_from_string(ref)
    c1 = column_index_from_string(cs)
    if to:
        ce, re_ = coordinate_from_string(to)
        ws.merge_cells(start_row=ri, start_column=c1, end_row=re_, end_column=column_index_from_string(ce))
    cc = ws.cell(row=ri, column=c1)
    cc.value = text
    cc.font = Font(name=FONT, size=size, bold=bold)
    cc.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    return cc

# ===== タイトル =====
cell('A1', '羽田定額適用外理由書', size=19, bold=True, ha='center', to=L(NCOL) + '1')

# ===== 営業所（表の外・PDF準拠で上部） =====
cell('B3', '営業所', size=11, ha='center', to='F3')
cell('H3', '', to='AB3')             # 営業所名（自動入力）
edge(3, 8, 3, 28, 'B')               # 下線

# ===== 課 / 班 / 20 年 月 日 =====
cell('B5', '課', size=11, ha='center', to='C5')
cell('D5', '', to='J5'); edge(5, 4, 5, 10, 'B')     # 課番号（自動）
cell('L5', '班', size=11, ha='center', to='M5')
cell('N5', '', to='S5'); edge(5, 14, 5, 19, 'B')    # 班（自動）
cell('U5', '20', size=11, ha='right')
cell('W5', '', to='Z5'); edge(5, 23, 5, 26, 'B')    # 年
cell('AA5', '年', size=11)
cell('AB5', '', to='AD5'); edge(5, 28, 5, 30, 'B')  # 月
cell('AE5', '月', size=11)
cell('AF5', '', to='AH5'); edge(5, 32, 5, 34, 'B')  # 日
cell('AI5', '日', size=11)

# ===== 明細テーブル（PDFの6行グリッド） 行7〜12 =====
grid(7, 1, 12, NCOL)

def vsplit(r1, r2, c):   # c列の左に縦罫（= c-1の右）
    for r in range(r1, r2 + 1):
        a = ws.cell(row=r, column=c); ab = a.border
        a.border = Border(top=ab.top, bottom=ab.bottom, left=S(), right=ab.right)
        p = ws.cell(row=r, column=c - 1); pb = p.border
        p.border = Border(top=pb.top, bottom=pb.bottom, left=pb.left, right=S())

def hsplit(r, c1, c2):   # r行の下に横罫
    for c in range(c1, c2 + 1):
        a = ws.cell(row=r, column=c); ab = a.border
        a.border = Border(top=ab.top, bottom=S(), left=ab.left, right=ab.right)

LC = 7          # ラベル列の右端（データ開始は col 8）
MIDC = 17       # 乗車地/降車地ラベルの左端
RC = 27         # 右ブロック（車両番号 / 乗車人数）の左端
RC2 = 33        # 「人」「可・不可」列の左端

vsplit(7, 12, LC + 1)         # ラベル列 | データ
vsplit(9, 12, MIDC)           # 時刻 | 乗車地
vsplit(7, 8, RC)              # 行7-8右: 氏名/車両番号 の左
vsplit(7, 12, RC)             # 行9-12右: 乗車地値 の左（= 349.0 相当）
vsplit(11, 12, RC2)           # 乗車人数 | 人
vsplit(7, 8, 25)             # 行7: 社員コード領域 | 氏名ラベル (270.8相当)
for r in (8, 9, 10, 11):
    hsplit(r, 1, NCOL)
hsplit(12, 8, NCOL)          # 乗客情報 内部横罫（左ラベル列は縦に貫通）

# 行7: 社員コード / 氏名
cell('B7', '社員コード', size=10.5, ha='center', to='G7')
cell('I7', '', to='X7')                 # 社員コード（自動）
cell('Y7', '氏名', size=10.5, ha='center', to='AB7')
cell('AC7', '', to=L(NCOL) + '7')       # 氏名（自動）

# 行8: 乗車日 / 20 年 月 日( ) / 車両番号
cell('B8', '乗車日', size=10.5, ha='center', to='G8')
cell('I8', '20', size=10.5, ha='right')
cell('K8', '', to='M8'); cell('N8', '年', size=10.5)
cell('O8', '', to='P8'); cell('Q8', '月', size=10.5)
cell('R8', '', to='S8'); cell('T8', '日（', size=10.5)
cell('U8', '', to='W8'); cell('X8', '）', size=10.5)
cell('AA8', '車両番号', size=10.5, ha='center', to='AF8')
cell('AG8', '', to=L(NCOL) + '8')       # 車両番号（入力）

# 行9: 乗車時間 : / 乗車地 / 羽田空港 T1・T2・T3
cell('B9', '乗車時間', size=10.5, ha='center', to='G9')
cell('J9', '', to='L9'); cell('M9', '：', size=10.5, ha='center')
cell('N9', '', to='P9')
cell('R9', '乗車地', size=10.5, ha='center', to='V9')
cell('AA9', '羽田空港    T1 ・ T2 ・ T3', size=10.5, ha='center', to=L(NCOL) + '9')

# 行10: 降車時間 : / 降車地 / (入力)
cell('B10', '降車時間', size=10.5, ha='center', to='G10')
cell('J10', '', to='L10'); cell('M10', '：', size=10.5, ha='center')
cell('N10', '', to='P10')
cell('R10', '降車地', size=10.5, ha='center', to='V10')
cell('AA10', '', to=L(NCOL) + '10')    # 降車地（入力）

# 行11-12: 乗客情報 / 無線配車・定額乗場 / 乗車人数 人
ws.merge_cells('A11:G12')
cell('A11', '乗客情報', size=10.5, ha='center', va='center')
cell('I11', '無線配車　・　定額乗場', size=10.5, ha='center', to='Z11')
cell('AA11', '乗車人数', size=10.5, ha='center', to='AF11')
cell('AG11', '', to='AL11')            # 乗車人数（入力）
cell('AM11', '人', size=10.5, ha='center', to=L(NCOL) + '11')
cell('I12', '日本人　・　外国人　・　不明', size=10.5, ha='center', to='Z12')
cell('AA12', '日本語', size=10.5, ha='center', to='AF12')
cell('AG12', '可　・　不可', size=10.5, ha='center', to=L(NCOL) + '12')

# ===== 適用外理由 =====
cell('A14', '適用外理由', size=12, bold=True, va='center', to='F14')
cell('H14', '※該当にチェック', size=10.5, va='center', to='R14')

cell('C16', '□', size=13, ha='center'); cell('E16', '高速道路利用なし', size=11, to='AB16')
cell('C17', '□', size=13, ha='center'); cell('E17', '羽田タクシー乗場以外から乗車', size=11, to='AB17')
cell('C18', '□', size=13, ha='center'); cell('E18', '経由地あり（定額範囲外）【経由地：', size=11, to='V18')
cell('W18', '', to='AL18')            # 経由地（入力）
cell('AM18', '】', size=11)
cell('C19', '□', size=13, ha='center'); cell('E19', 'その他', size=11, to='AB19')

# その他 記入用 点線罫
for r in (20, 21, 22, 23):
    for c in range(4, NCOL + 1):
        a = ws.cell(row=r, column=c); ab = a.border
        a.border = Border(top=ab.top, bottom=S('dotted'), left=ab.left, right=ab.right)
for r in (20, 21, 22, 23):
    cell('D%d' % r, '', to=L(NCOL) + str(r))

# ===== 右下 課長 / 確認者 =====
grid(28, 27, 29, NCOL)
vsplit(28, 29, 34)
hsplit(28, 27, NCOL)
cell('AA28', '課長', size=10.5, ha='center', to='AG28')
cell('AH28', '確認者', size=10.5, ha='center', to=L(NCOL) + '28')
cell('AA29', '', size=11, ha='center', va='center', to='AG29')      # 課長名（マスタ）
cell('AH29', '', size=11, ha='center', va='center', to=L(NCOL) + '29')   # 確認者名

# ===== 印刷設定 =====
ws.page_setup.orientation = 'portrait'
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = 0.6
ws.page_margins.right = 0.5
ws.page_margins.top = 0.6
ws.page_margins.bottom = 0.4
ws.page_margins.header = 0.3
ws.page_margins.footer = 0.3
ws.print_options.horizontalCentered = True

# ===== fill 対象セルを確実に出力 =====
FILL_REFS = [
    'H3',                      # 営業所名
    'D5', 'N5',                # 課 / 班
    'W5', 'AB5', 'AF5',        # 年 / 月 / 日（作成日）
    'I7', 'AC7',               # 社員コード / 氏名
    'K8', 'O8', 'R8', 'U8',    # 乗車日 年/月/日/曜日
    'AG8',                     # 車両番号
    'J9', 'N9',                # 乗車時間 時/分
    'J10', 'N10',              # 降車時間 時/分
    'AA9',                     # 乗車地（既定テキスト上書き）
    'AA10',                    # 降車地
    'AG11',                    # 乗車人数
    'W18',                     # 経由地
    'D20', 'D21', 'D22', 'D23',  # その他 記入行
    'AA29', 'AH29',            # 課長 / 確認者
]
for ref in FILL_REFS:
    cs, ri = coordinate_from_string(ref)
    ci = column_index_from_string(cs)
    cc = ws.cell(row=ri, column=ci)
    if isinstance(cc, MergedCell):
        raise SystemExit('FILL ref %s is a non-anchor MergedCell' % ref)
    if cc.value is None:
        cc.value = ''
    if cc.font is None or cc.font.name is None:
        cc.font = Font(name=FONT, size=10.5)
    if cc.alignment is None or cc.alignment.horizontal is None:
        cc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

wb.save(OUT_XLSX)
raw = open(OUT_XLSX, 'rb').read()
b64 = base64.b64encode(raw).decode()
with open(OUT_TS, 'w') as fp:
    fp.write('// 羽田定額適用外理由書（国際自動車）のひな型。PDF フラット帳票を xlsx で再現。\n')
    fp.write('export const HANEDA_RIYUSHO_TEMPLATE_XLSX_B64 =\n')
    parts = [b64[i:i + 100] for i in range(0, len(b64), 100)]
    fp.write('  "' + '" +\n  "'.join(parts) + '";\n')

z = zipfile.ZipFile(io.BytesIO(raw))
s1 = z.read('xl/worksheets/sheet1.xml').decode()
missing = [ref for ref in FILL_REFS if ('r="%s"' % ref) not in s1]
print('xlsx bytes', len(raw), 'missing fill refs:', missing, 'merges:', s1.count('<mergeCell '))
