# 顛末書.xls (旧BIFF) を xlrd で読み、罫線・結合・列幅・行高・フォント・配置を
# そのまま openpyxl の xlsx に再現する。値の差し替えはブラウザ側 (xlsx_fill_client) で行う。
import base64, io, sys
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = '/Users/reito/Downloads/顛末書.xls'
OUT_XLSX = 'tenmatsusho_template.xlsx'
OUT_TS = 'tenmatsusho_template.ts'

LS = {1: 'thin', 2: 'medium', 3: 'dashed', 4: 'dotted', 5: 'thick', 6: 'double', 7: 'hair'}
HA = {1: 'left', 2: 'center', 3: 'right', 4: 'fill', 5: 'justify', 6: 'centerContinuous', 7: 'distributed'}
VA = {0: 'top', 1: 'center', 2: 'bottom', 3: 'justify', 4: 'distributed'}

b = xlrd.open_workbook(SRC, formatting_info=True)
sh = b.sheet_by_index(0)

wb = Workbook()
ws = wb.active
ws.title = '顛末書'

# 列幅
for c in range(sh.ncols):
    ci = sh.colinfo_map.get(c)
    w = (ci.width / 256.0) if ci else (sh.defcolwidth or 8.43)
    ws.column_dimensions[get_column_letter(c + 1)].width = round(w, 3)

# 行高
for r in range(sh.nrows):
    ri = sh.rowinfo_map.get(r)
    if ri and ri.height:
        ws.row_dimensions[r + 1].height = round(ri.height / 20.0, 2)

def side(style_idx):
    st = LS.get(style_idx)
    return Side(style=st) if st else None

# セル書式・値
for r in range(sh.nrows):
    for c in range(sh.ncols):
        xfi = sh.cell_xf_index(r, c)
        xf = b.xf_list[xfi]
        f = b.font_list[xf.font_index]
        bd = xf.border
        al = xf.alignment
        has_border = any([bd.top_line_style, bd.bottom_line_style, bd.left_line_style, bd.right_line_style])
        v = sh.cell_value(r, c)
        # 罫線が無く値も無いセルは省略（ただし後で fill 対象を強制生成）
        if not has_border and (v == '' or v is None):
            continue
        cell = ws.cell(row=r + 1, column=c + 1)
        if v != '' and v is not None:
            cell.value = v
        cell.font = Font(
            name=f.name or 'ＭＳ Ｐ明朝',
            size=(f.height / 20.0) if f.height else 11,
            bold=bool(f.bold),
        )
        cell.alignment = Alignment(
            horizontal=HA.get(al.hor_align),
            vertical=VA.get(al.vert_align, 'center'),
            wrap_text=bool(al.text_wrapped),
        )
        if has_border:
            cell.border = Border(
                top=side(bd.top_line_style), bottom=side(bd.bottom_line_style),
                left=side(bd.left_line_style), right=side(bd.right_line_style),
            )

# 結合セル
for (rlo, rhi, clo, chi) in sh.merged_cells:
    ws.merge_cells(start_row=rlo + 1, start_column=clo + 1, end_row=rhi, end_column=chi)

# ---- 差し替え対象セルを強制生成（openpyxl は空セルを書き出さないため） ----
# 参照は build 後に grep で確認できるようにする
FILL_REFS = [
    'P2',   # 提出年（右上 "年" は R2）
    'S2',   # 提出月（"月" は T2）
    'U2',   # 提出日（"日" は V2）
    'P5',   # 課（"課" は S5）
    'T5',   # 班（"班" は V5）
    'P6',   # 氏名（P6:U6 結合の先頭）
    'P7',   # コード（P7:V7 結合の先頭）
    'E13',  # 発生年（E13:F13 結合の先頭）
    'H13',  # 発生月
    'J13',  # 発生日
    'M13',  # 曜日
    'R13',  # 発生時
    'T13',  # 発生分
    'E14',  # 発生場所（E14:V14 結合の先頭）
    'E15',  # 件名（E15:O15 結合の先頭）
    'S15',  # コールサイン（S15:V15 結合の先頭）
    'A20',  # 本文（記）: 事前印字文の下段 A20:V20
    'A32',  # 所見および処置等（A32:R32 結合の先頭）
    'S32',  # 事業所責任者（S32:T33 結合の先頭）
    'U32',  # 運行管理者（U32:V33 結合の先頭）
]
thin_all = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.cell.cell import MergedCell
for ref in FILL_REFS:
    col_s, row_i = coordinate_from_string(ref)
    ci = column_index_from_string(col_s)
    cell = ws.cell(row=row_i, column=ci)
    if isinstance(cell, MergedCell):
        raise SystemExit('FILL ref %s is a non-anchor MergedCell — fix mapping' % ref)
    if cell.value is None:
        cell.value = ''
    # 既存の書式を保持。無ければ標準フォントだけ付けて <c> を確実に出力させる
    if cell.font is None or cell.font.name is None:
        cell.font = Font(name='ＭＳ Ｐ明朝', size=11)
    if not (cell.alignment and cell.alignment.horizontal):
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 印刷設定 A4縦・1ページ幅
ws.page_setup.orientation = 'portrait'
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = 0.51
ws.page_margins.right = 0.31
ws.page_margins.top = 0.55
ws.page_margins.bottom = 0.35
ws.page_margins.header = 0.3
ws.page_margins.footer = 0.3

wb.save(OUT_XLSX)

raw = open(OUT_XLSX, 'rb').read()
b64 = base64.b64encode(raw).decode()
with open(OUT_TS, 'w') as fp:
    fp.write('// 顛末書のひな型（国際自動車株式会社 T2 / 旧 xls を xlsx 再現）。書式・結合・印刷設定はそのまま。\n')
    fp.write('export const TENMATSUSHO_TEMPLATE_XLSX_B64 =\n')
    # 100文字ごとに改行して JS 文字列連結
    step = 100
    parts = [b64[i:i+step] for i in range(0, len(b64), step)]
    fp.write('  "' + '" +\n  "'.join(parts) + '";\n')

print('xlsx bytes', len(raw), ' b64 len', len(b64))
# 検証: sheet1.xml に FILL_REFS が入っているか
import zipfile
z = zipfile.ZipFile(io.BytesIO(raw))
s1 = z.read('xl/worksheets/sheet1.xml').decode()
missing = [ref for ref in FILL_REFS if ('r="%s"' % ref) not in s1]
print('missing fill refs:', missing)
print('merged count:', s1.count('<mergeCell '))
