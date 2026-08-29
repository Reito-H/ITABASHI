import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import range_boundaries

path = sys.argv[1]
wb = load_workbook(path)
ws = wb.active
maxc = ws.max_column
maxr = ws.max_row

# merged map: anchor -> (rowspan, colspan); covered set
covered = set()
spans = {}
for mr in ws.merged_cells.ranges:
    c1, r1, c2, r2 = range_boundaries(str(mr))
    spans[(r1, c1)] = (r2 - r1 + 1, c2 - c1 + 1)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            if (r, c) != (r1, c1):
                covered.add((r, c))

def px_w(c):
    w = ws.column_dimensions[get_column_letter(c)].width or 8.43
    return round(w * 7 + 5)
def px_h(r):
    h = ws.row_dimensions[r].height or 15
    return round(h * 1.33)

out = ['<!doctype html><meta charset=utf-8><style>',
       'table{border-collapse:collapse;font-family:"MS PMincho",serif}',
       'td{overflow:hidden;font-size:11px;padding:0 1px;vertical-align:middle}',
       '</style><table>']
# colgroup
out.append('<colgroup>' + ''.join(f'<col style="width:{px_w(c)}px">' for c in range(1, maxc + 1)) + '</colgroup>')
for r in range(1, maxr + 1):
    out.append(f'<tr style="height:{px_h(r)}px">')
    for c in range(1, maxc + 1):
        if (r, c) in covered:
            continue
        cell = ws.cell(row=r, column=c)
        rs, cs = spans.get((r, c), (1, 1))
        b = cell.border
        st = []
        for side, name in [(b.top, 'top'), (b.bottom, 'bottom'), (b.left, 'left'), (b.right, 'right')]:
            if side and side.style:
                w = '2px' if side.style in ('medium', 'thick') else '1px'
                sty = 'dotted' if side.style in ('dotted', 'dashed', 'hair') else 'solid'
                st.append(f'border-{name}:{w} {sty} #000')
        al = cell.alignment
        if al and al.horizontal:
            st.append('text-align:' + (al.horizontal if al.horizontal in ('left','right','center') else 'left'))
        if cell.font and cell.font.size:
            st.append(f'font-size:{round(cell.font.size)}px')
        if cell.font and cell.font.bold:
            st.append('font-weight:700')
        v = '' if cell.value is None else str(cell.value)
        out.append(f'<td rowspan={rs} colspan={cs} style="{";".join(st)}">{v}</td>')
    out.append('</tr>')
out.append('</table>')
open(sys.argv[2], 'w').write('\n'.join(out))
print('wrote', sys.argv[2])
